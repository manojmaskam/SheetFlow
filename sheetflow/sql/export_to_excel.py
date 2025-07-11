import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from datetime import datetime
from sheetflow.db.connection_pool import get_connection
from sheetflow.config import get_config
import logging

# Setup logs
log_filename = f"logs/export_{datetime.now().strftime('%Y-%m-%d')}.log"
logging.basicConfig(
    filename=log_filename,
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

def export_sql_to_excel():
    config = get_config()
    table = config.get("DATABASE", "table")
    excel_path = config.get("EXCEL", "file_path")

    # ✅ Connect to DB and fetch data
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute(f"SELECT * FROM {table}")
    data = cursor.fetchall()
    conn.close()
    cursor.close()

    # ✅ Create DataFrame without column titles
    df = pd.DataFrame(data)

    # ✅ Load existing Excel and write below existing headers (row 2 onwards)
    workbook = load_workbook(excel_path)
    sheet = workbook.active

    for r_idx, row in enumerate(data, start=2):  # start at row 2
        for c_idx, value in enumerate(row, start=1):
            cell = sheet.cell(row=r_idx, column=c_idx, value=value)

            # Format date column (assuming 3rd column is Appointment Date)
            if c_idx == 3:
                cell.number_format = "DD-MM-YYYY"
                cell.alignment = Alignment(horizontal='center')

    workbook.save(excel_path)

    print(f"📤 Exported {len(df)} rows to Excel without altering header → {excel_path}")
    logging.info(f"📤 Exported {len(df)} rows to Excel without altering header → {excel_path}")

if __name__ == "__main__":
    export_sql_to_excel()
